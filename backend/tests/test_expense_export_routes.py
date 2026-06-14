"""
Route tests for GET /api/expenses/export.

Uses smart_receipt_db_test — never touches the development DB.
"""

from datetime import date, datetime
from decimal import Decimal
from io import BytesIO

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from app.config import settings
from app.database import get_db
from app.main import app
from app.models.category import Category
from app.models.expense import Expense
from app.models.expense_item import ExpenseItem
from app.models.receipt_file import ReceiptFile
from app.models.refresh_token import RefreshToken
from app.models.user import User

# ── Test DB setup ──────────────────────────────────────────────────────────────

settings.JWT_SECRET_KEY = "test_jwt_secret_key_export_routes_testing_2026"
settings.JWT_ALGORITHM = "HS256"

engine = create_engine(settings.TEST_DATABASE_URL)
TestingSessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)


def override_get_db():
    db = TestingSessionLocal()
    try:
        yield db
    finally:
        db.close()


app.dependency_overrides[get_db] = override_get_db
client = TestClient(app)


# ── Clean-up fixture ───────────────────────────────────────────────────────────

@pytest.fixture(autouse=True)
def clean_db():
    db = TestingSessionLocal()
    db.query(ExpenseItem).delete()
    db.query(ReceiptFile).delete()
    db.query(Expense).delete()
    db.query(Category).delete()
    db.query(RefreshToken).delete()
    db.query(User).delete()
    db.commit()
    yield
    db.query(ExpenseItem).delete()
    db.query(ReceiptFile).delete()
    db.query(Expense).delete()
    db.query(Category).delete()
    db.query(RefreshToken).delete()
    db.query(User).delete()
    db.commit()
    db.close()


# ── Helpers ────────────────────────────────────────────────────────────────────

def register_and_login(username="exp_export", email="exp_export@example.com") -> str:
    client.post("/api/auth/register", json={
        "username": username,
        "email": email,
        "password": "password123",
    })
    res = client.post("/api/auth/login", json={
        "email": email,
        "password": "password123",
    })
    return res.json()["access_token"]


def make_category(code="FOOD", name_en="Food", name_th="อาหาร") -> int:
    db = TestingSessionLocal()
    cat = Category(code=code, name_en=name_en, name_th=name_th, is_active=True)
    db.add(cat)
    db.commit()
    db.refresh(cat)
    cat_id = cat.id
    db.close()
    return cat_id


def create_expense_api(token: str, category_id: int, title: str = "Expense",
                       items: list = None) -> dict:
    payload = {
        "category_id": category_id,
        "title": title,
        "receipt_date": "2026-01-15",
        "total_amount": "100.00",
        "currency": "THB",
        "items": items or [],
    }
    res = client.post(
        "/api/expenses",
        json=payload,
        headers={"Authorization": f"Bearer {token}"},
    )
    assert res.status_code == 201, f"Setup failed: {res.json()}"
    return res.json()


def soft_delete_expense(expense_id: int) -> None:
    db = TestingSessionLocal()
    e = db.query(Expense).filter(Expense.id == expense_id).first()
    if e:
        e.deleted_at = datetime(2024, 1, 1)
        db.commit()
    db.close()


def do_export(token: str):
    return client.get(
        "/api/expenses/export",
        headers={"Authorization": f"Bearer {token}"},
    )


def load_workbook_from_response(res) -> object:
    return load_workbook(BytesIO(res.content))


# ── Fixtures ───────────────────────────────────────────────────────────────────

@pytest.fixture()
def token():
    return register_and_login()


@pytest.fixture()
def other_token():
    return register_and_login("other_exp", "other_exp@example.com")


@pytest.fixture()
def category_id():
    return make_category()


# ── Tests ──────────────────────────────────────────────────────────────────────

def test_authenticated_user_can_export(token):
    res = do_export(token)
    assert res.status_code == 200


def test_export_returns_200(token):
    res = do_export(token)
    assert res.status_code == 200


def test_response_mime_type_is_excel(token):
    res = do_export(token)
    assert "spreadsheetml" in res.headers["content-type"]


def test_content_disposition_contains_xlsx(token):
    res = do_export(token)
    cd = res.headers.get("content-disposition", "")
    assert ".xlsx" in cd


def test_response_body_is_valid_excel_workbook(token):
    res = do_export(token)
    wb = load_workbook_from_response(res)
    assert wb is not None


def test_workbook_contains_both_worksheets(token):
    res = do_export(token)
    wb = load_workbook_from_response(res)
    assert "Expenses" in wb.sheetnames
    assert "Expense Items" in wb.sheetnames


def test_only_current_user_data_appears(token, other_token, category_id):
    create_expense_api(token, category_id, title="Mine")
    create_expense_api(other_token, category_id, title="Theirs")

    res = do_export(token)
    wb = load_workbook_from_response(res)
    ws = wb["Expenses"]
    titles = [ws.cell(row=r, column=4).value for r in range(2, ws.max_row + 1)]
    assert "Mine" in titles
    assert "Theirs" not in titles


def test_another_users_data_does_not_appear(token, other_token, category_id):
    create_expense_api(other_token, category_id, title="Not Mine At All")

    res = do_export(token)
    wb = load_workbook_from_response(res)
    ws = wb["Expenses"]
    assert ws.max_row == 1  # header only


def test_soft_deleted_data_does_not_appear(token, category_id):
    active = create_expense_api(token, category_id, title="Active")
    deleted = create_expense_api(token, category_id, title="Deleted")
    soft_delete_expense(deleted["id"])

    res = do_export(token)
    wb = load_workbook_from_response(res)
    ws = wb["Expenses"]
    titles = [ws.cell(row=r, column=4).value for r in range(2, ws.max_row + 1)]
    assert "Active" in titles
    assert "Deleted" not in titles


def test_empty_export_returns_valid_workbook(token):
    res = do_export(token)
    assert res.status_code == 200
    wb = load_workbook_from_response(res)
    ws = wb["Expenses"]
    assert ws.max_row == 1  # header only


def test_missing_token_returns_401():
    res = client.get("/api/expenses/export")
    assert res.status_code == 401


def test_invalid_token_returns_401():
    res = client.get(
        "/api/expenses/export",
        headers={"Authorization": "Bearer bad.token.here"},
    )
    assert res.status_code == 401


def test_export_is_not_treated_as_expense_id(token):
    """
    'export' must resolve to the export route, not /{expense_id}.
    If it were treated as an ID it would 404 (no expense with id='export').
    The export route returns 200 even for an empty export.
    """
    res = do_export(token)
    assert res.status_code == 200


def test_sensitive_fields_not_in_workbook(token, category_id):
    create_expense_api(token, category_id, title="Check Fields")

    res = do_export(token)
    wb = load_workbook_from_response(res)
    ws = wb["Expenses"]

    headers = [cell.value for cell in ws[1]]
    for forbidden in ("ai_raw_response", "deleted_at", "user_id", "ai_confidence",
                      "ai_status", "password", "token"):
        assert forbidden not in headers, f"Sensitive header found: {forbidden}"


def test_expense_items_appear_in_export(token, category_id):
    items = [
        {"original_name": "Coffee", "quantity": "1", "total_price": "50.00"},
        {"original_name": "Cake", "quantity": "2", "total_price": "80.00"},
    ]
    create_expense_api(token, category_id, items=items)

    res = do_export(token)
    wb = load_workbook_from_response(res)
    ws = wb["Expense Items"]
    names = [ws.cell(row=r, column=3).value for r in range(2, ws.max_row + 1)]
    assert "Coffee" in names
    assert "Cake" in names


def test_content_disposition_contains_today_date(token):
    res = do_export(token)
    cd = res.headers.get("content-disposition", "")
    today = date.today().isoformat()
    assert today in cd


def test_existing_full_test_suite_smoke(token, category_id):
    """Smoke: existing create / list / detail routes still work alongside export."""
    expense = create_expense_api(token, category_id, title="Smoke")

    detail_res = client.get(
        f"/api/expenses/{expense['id']}",
        headers={"Authorization": f"Bearer {token}"},
    )
    assert detail_res.status_code == 200

    list_res = client.get("/api/expenses", headers={"Authorization": f"Bearer {token}"})
    assert list_res.status_code == 200

    export_res = do_export(token)
    assert export_res.status_code == 200
