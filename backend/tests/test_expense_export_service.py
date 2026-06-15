"""
Service-layer tests for export_user_expenses_to_excel.

Uses smart_receipt_db_test — never touches the development DB.
"""

from datetime import date, datetime
from decimal import Decimal
from io import BytesIO

import pytest
from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from app.config import settings
from app.models.category import Category
from app.models.expense import Expense
from app.models.expense_item import ExpenseItem
from app.models.refresh_token import RefreshToken
from app.models.user import User
from app.services.expense_service import export_user_expenses_to_excel

# ── Test DB setup ──────────────────────────────────────────────────────────────

engine = create_engine(settings.TEST_DATABASE_URL)
TestingSessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)


@pytest.fixture(autouse=True)
def clean_db():
    db = TestingSessionLocal()
    db.query(ExpenseItem).delete()
    db.query(Expense).delete()
    db.query(Category).delete()
    db.query(RefreshToken).delete()
    db.query(User).delete()
    db.commit()
    yield
    db.query(ExpenseItem).delete()
    db.query(Expense).delete()
    db.query(Category).delete()
    db.query(RefreshToken).delete()
    db.query(User).delete()
    db.commit()
    db.close()


# ── Helpers ────────────────────────────────────────────────────────────────────

def make_user(username="tester", email="tester@example.com") -> User:
    db = TestingSessionLocal()
    user = User(
        username=username,
        email=email,
        password_hash="fakehash",
        is_active=True,
    )
    db.add(user)
    db.commit()
    db.refresh(user)
    db.close()
    return user


def make_category(code="FOOD", name_en="Food", name_th="อาหาร") -> Category:
    db = TestingSessionLocal()
    cat = Category(code=code, name_en=name_en, name_th=name_th, is_active=True)
    db.add(cat)
    db.commit()
    db.refresh(cat)
    db.close()
    return cat


def make_expense(
    user_id: int,
    category_id=None,
    title="Lunch",
    total_amount=Decimal("100.00"),
    deleted_at=None,
    input_method="manual",
) -> Expense:
    db = TestingSessionLocal()
    exp = Expense(
        user_id=user_id,
        category_id=category_id,
        title=title,
        receipt_date=date(2026, 1, 15),
        total_amount=total_amount,
        currency="THB",
        input_method=input_method,
        document_type="manual_expense",
        deleted_at=deleted_at,
    )
    db.add(exp)
    db.commit()
    db.refresh(exp)
    db.close()
    return exp


def make_item(
    expense_id: int,
    category_id=None,
    original_name="Item",
    name_en=None,
    name_th=None,
    quantity=Decimal("1.000"),
    total_price=Decimal("50.00"),
    deleted_at=None,
) -> ExpenseItem:
    db = TestingSessionLocal()
    item = ExpenseItem(
        expense_id=expense_id,
        category_id=category_id,
        original_name=original_name,
        name_en=name_en,
        name_th=name_th,
        quantity=quantity,
        total_price=total_price,
        discount_amount=Decimal("0.00"),
        deleted_at=deleted_at,
    )
    db.add(item)
    db.commit()
    db.refresh(item)
    db.close()
    return item


def run_export(user_id: int) -> BytesIO:
    db = TestingSessionLocal()
    try:
        return export_user_expenses_to_excel(db, user_id)
    finally:
        db.close()


def open_workbook(stream: BytesIO) -> object:
    """Load the workbook from a BytesIO stream."""
    stream.seek(0)
    return load_workbook(stream)


# ── Tests ──────────────────────────────────────────────────────────────────────

def test_export_returns_bytesio():
    user = make_user()
    result = run_export(user.id)
    assert isinstance(result, BytesIO)


def test_returned_workbook_can_be_opened_by_openpyxl():
    user = make_user()
    stream = run_export(user.id)
    wb = open_workbook(stream)
    assert wb is not None


def test_workbook_contains_expenses_sheet():
    user = make_user()
    stream = run_export(user.id)
    wb = open_workbook(stream)
    assert "Expenses" in wb.sheetnames


def test_workbook_contains_expense_items_sheet():
    user = make_user()
    stream = run_export(user.id)
    wb = open_workbook(stream)
    assert "Expense Items" in wb.sheetnames


def test_expenses_sheet_has_correct_headers():
    user = make_user()
    stream = run_export(user.id)
    wb = open_workbook(stream)
    ws = wb["Expenses"]
    headers = [cell.value for cell in ws[1]]
    assert "Expense ID" in headers
    assert "Title" in headers
    assert "Total Amount" in headers
    assert "Category" in headers
    assert "Confirmed" in headers
    assert "Receipt Date" in headers
    assert "Created At" in headers


def test_expense_items_sheet_has_correct_headers():
    user = make_user()
    stream = run_export(user.id)
    wb = open_workbook(stream)
    ws = wb["Expense Items"]
    headers = [cell.value for cell in ws[1]]
    assert "Expense ID" in headers
    assert "Item ID" in headers
    assert "Original Name" in headers
    assert "Total Price" in headers
    assert "Category" in headers


def test_only_current_users_expenses_are_exported():
    user_a = make_user(username="a", email="a@x.com")
    user_b = make_user(username="b", email="b@x.com")
    cat = make_category()
    make_expense(user_a.id, cat.id, title="A's Expense")
    make_expense(user_b.id, cat.id, title="B's Expense")

    stream = run_export(user_a.id)
    wb = open_workbook(stream)
    ws = wb["Expenses"]
    titles = [ws.cell(row=r, column=4).value for r in range(2, ws.max_row + 1)]
    assert "A's Expense" in titles
    assert "B's Expense" not in titles


def test_another_users_expenses_are_excluded():
    user_a = make_user(username="a2", email="a2@x.com")
    user_b = make_user(username="b2", email="b2@x.com")
    cat = make_category(code="FOOD2", name_en="Food2", name_th="อาหาร2")
    make_expense(user_b.id, cat.id, title="Not Mine")

    stream = run_export(user_a.id)
    wb = open_workbook(stream)
    ws = wb["Expenses"]
    assert ws.max_row == 1  # only the header row


def test_soft_deleted_expenses_are_excluded():
    user = make_user(username="del_u", email="del@x.com")
    cat = make_category(code="DEL", name_en="Del", name_th="ลบ")
    make_expense(user.id, cat.id, title="Active")
    make_expense(user.id, cat.id, title="Deleted", deleted_at=datetime(2024, 1, 1))

    stream = run_export(user.id)
    wb = open_workbook(stream)
    ws = wb["Expenses"]
    titles = [ws.cell(row=r, column=4).value for r in range(2, ws.max_row + 1)]
    assert "Active" in titles
    assert "Deleted" not in titles


def test_soft_deleted_items_are_excluded():
    user = make_user(username="item_del", email="itemdel@x.com")
    cat = make_category(code="ITEM_D", name_en="ItemDel", name_th="ลบ")
    exp = make_expense(user.id, cat.id)
    make_item(exp.id, original_name="Visible")
    make_item(exp.id, original_name="Gone", deleted_at=datetime(2024, 1, 1))

    stream = run_export(user.id)
    wb = open_workbook(stream)
    ws = wb["Expense Items"]
    names = [ws.cell(row=r, column=3).value for r in range(2, ws.max_row + 1)]
    assert "Visible" in names
    assert "Gone" not in names


def test_active_items_are_included():
    user = make_user(username="active_u", email="active@x.com")
    cat = make_category(code="ACT", name_en="Active", name_th="ใช้งาน")
    exp = make_expense(user.id, cat.id)
    make_item(exp.id, original_name="Coffee")
    make_item(exp.id, original_name="Tea")

    stream = run_export(user.id)
    wb = open_workbook(stream)
    ws = wb["Expense Items"]
    names = [ws.cell(row=r, column=3).value for r in range(2, ws.max_row + 1)]
    assert "Coffee" in names
    assert "Tea" in names


def test_null_expense_category_displays_uncategorized():
    user = make_user(username="nocat_u", email="nocat@x.com")
    make_expense(user.id, category_id=None, title="No Category")

    stream = run_export(user.id)
    wb = open_workbook(stream)
    ws = wb["Expenses"]
    # Column 6 = Category
    cat_value = ws.cell(row=2, column=6).value
    assert cat_value == "Uncategorized"


def test_null_item_category_displays_uncategorized():
    user = make_user(username="nocat_i", email="nocati@x.com")
    cat = make_category(code="MAIN2", name_en="Main2", name_th="หลัก2")
    exp = make_expense(user.id, cat.id)
    make_item(exp.id, category_id=None, original_name="No-cat item")

    stream = run_export(user.id)
    wb = open_workbook(stream)
    ws = wb["Expense Items"]
    # Column 6 = Category
    cat_value = ws.cell(row=2, column=6).value
    assert cat_value == "Uncategorized"


def test_category_name_is_readable_english():
    user = make_user(username="cat_en", email="caten@x.com")
    cat = make_category(code="TRAVEL", name_en="Travel", name_th="ท่องเที่ยว")
    make_expense(user.id, cat.id, title="Trip")

    stream = run_export(user.id)
    wb = open_workbook(stream)
    ws = wb["Expenses"]
    cat_value = ws.cell(row=2, column=6).value
    assert cat_value == "Travel"


def test_decimal_money_values_are_numeric():
    user = make_user(username="num_u", email="num@x.com")
    cat = make_category(code="NUM", name_en="Num", name_th="ตัวเลข")
    make_expense(user.id, cat.id, total_amount=Decimal("250.75"))

    stream = run_export(user.id)
    wb = open_workbook(stream)
    ws = wb["Expenses"]
    # Column 14 = Total Amount
    total_cell = ws.cell(row=2, column=14)
    assert isinstance(total_cell.value, (int, float))
    assert abs(total_cell.value - 250.75) < 0.001


def test_money_cells_use_two_decimal_format():
    user = make_user(username="fmt_u", email="fmt@x.com")
    cat = make_category(code="FMT", name_en="Fmt", name_th="รูปแบบ")
    make_expense(user.id, cat.id, total_amount=Decimal("99.00"))

    stream = run_export(user.id)
    wb = open_workbook(stream)
    ws = wb["Expenses"]
    total_cell = ws.cell(row=2, column=14)
    assert total_cell.number_format == "0.00"


def test_receipt_date_uses_date_format():
    user = make_user(username="date_u", email="date@x.com")
    cat = make_category(code="DATE", name_en="Date", name_th="วันที่")
    make_expense(user.id, cat.id)

    stream = run_export(user.id)
    wb = open_workbook(stream)
    ws = wb["Expenses"]
    # Column 2 = Receipt Date
    date_cell = ws.cell(row=2, column=2)
    assert date_cell.number_format == "yyyy-mm-dd"


def test_header_row_is_frozen():
    user = make_user(username="freeze_u", email="freeze@x.com")
    stream = run_export(user.id)
    wb = open_workbook(stream)
    ws = wb["Expenses"]
    assert ws.freeze_panes == "A2"


def test_auto_filter_exists_on_expenses_sheet():
    user = make_user(username="filter_u", email="filter@x.com")
    cat = make_category(code="FIL", name_en="Fil", name_th="กรอง")
    make_expense(user.id, cat.id)

    stream = run_export(user.id)
    wb = open_workbook(stream)
    ws = wb["Expenses"]
    assert ws.auto_filter.ref is not None


def test_empty_user_export_returns_valid_workbook_with_headers():
    user = make_user(username="empty_u", email="empty@x.com")
    stream = run_export(user.id)
    wb = open_workbook(stream)
    assert "Expenses" in wb.sheetnames
    assert "Expense Items" in wb.sheetnames

    ws_exp = wb["Expenses"]
    assert ws_exp.max_row == 1  # header only

    ws_items = wb["Expense Items"]
    assert ws_items.max_row == 1  # header only


def test_ai_raw_response_is_not_exported():
    user = make_user(username="ai_u", email="ai@x.com")
    cat = make_category(code="AI", name_en="AI", name_th="เอไอ")
    exp = make_expense(user.id, cat.id)

    # Inject ai_raw_response directly
    db = TestingSessionLocal()
    e = db.query(Expense).filter(Expense.id == exp.id).first()
    e.ai_raw_response = {"raw": "secret data"}
    db.commit()
    db.close()

    stream = run_export(user.id)
    wb = open_workbook(stream)
    ws = wb["Expenses"]

    all_values = []
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                all_values.append(str(cell.value))

    assert "ai_raw_response" not in all_values
    assert "secret data" not in " ".join(all_values)


def test_deleted_at_is_not_exported():
    user = make_user(username="dlt_u", email="dlt@x.com")
    cat = make_category(code="DLT", name_en="Dlt", name_th="ลบ2")
    make_expense(user.id, cat.id)

    stream = run_export(user.id)
    wb = open_workbook(stream)
    ws = wb["Expenses"]

    headers = [cell.value for cell in ws[1]]
    assert "deleted_at" not in headers
    assert "Deleted At" not in headers


def test_local_receipt_paths_are_not_exported():
    user = make_user(username="path_u", email="path@x.com")
    cat = make_category(code="PATH", name_en="Path", name_th="เส้นทาง")
    make_expense(user.id, cat.id)

    stream = run_export(user.id)
    wb = open_workbook(stream)
    ws = wb["Expenses"]

    headers = [cell.value for cell in ws[1]]
    # No file path columns should appear
    for h in headers:
        if h:
            assert "path" not in h.lower(), f"Found path-like header: {h}"
            assert "file" not in h.lower(), f"Found file-like header: {h}"


def test_stable_ordering_expenses_ascending_by_created_at():
    """Expenses appear in created_at ASC order."""
    import time
    user = make_user(username="ord_u", email="ord@x.com")
    cat = make_category(code="ORD", name_en="Ord", name_th="ลำดับ")
    make_expense(user.id, cat.id, title="First")
    time.sleep(0.05)
    make_expense(user.id, cat.id, title="Second")

    stream = run_export(user.id)
    wb = open_workbook(stream)
    ws = wb["Expenses"]
    titles = [ws.cell(row=r, column=4).value for r in range(2, ws.max_row + 1)]
    assert titles == ["First", "Second"]


def test_no_database_rows_created_or_modified():
    """Export must be read-only — expense count must not change."""
    user = make_user(username="ro_u", email="ro@x.com")
    cat = make_category(code="RO", name_en="Readonly", name_th="อ่านอย่างเดียว")
    make_expense(user.id, cat.id, title="Before Export")

    db = TestingSessionLocal()
    count_before = db.query(Expense).filter(Expense.user_id == user.id).count()
    db.close()

    run_export(user.id)

    db = TestingSessionLocal()
    count_after = db.query(Expense).filter(Expense.user_id == user.id).count()
    db.close()

    assert count_before == count_after
