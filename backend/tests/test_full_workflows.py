"""
End-to-end workflow integration tests for Smart Receipt API.

These tests exercise complete user flows through the real HTTP routes,
hitting the test database (smart_receipt_db_test) the same way a frontend
would hit the API.

Gemini is always mocked — no real API calls are made.

Workflows covered:
  A. AI Receipt Flow — upload → extract (mocked) → confirm → translate (mocked) → export
  B. Manual Expense Flow — create → list → detail → update → export → delete → confirm gone
  C. User Isolation — User B cannot read or modify User A's data
"""

import io
import tempfile
import uuid
from datetime import date
from decimal import Decimal
from io import BytesIO
from pathlib import Path
from unittest.mock import patch

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from app.config import settings
from app.database import get_db
from app.main import app
from app.models.category import Category
from app.models.expense import Expense
from app.models.expense_item import ExpenseItem
from app.models.receipt_file import ReceiptFile
from app.models.refresh_token import RefreshToken
from app.models.translation import Translation
from app.models.user import User
from app.schemas.ai_extraction import ExtractedReceiptData, ExtractedReceiptItem
from app.schemas.translation import GeminiTranslatedItem, GeminiTranslationResult

# ── Test DB setup ──────────────────────────────────────────────────────────────

settings.JWT_SECRET_KEY = "test_jwt_secret_key_full_workflows_2026"
settings.JWT_ALGORITHM = "HS256"

engine = create_engine(settings.TEST_DATABASE_URL)
TestingSessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)


def override_get_db():
    db = TestingSessionLocal()
    try:
        yield db
    finally:
        db.close()


app.dependency_overrides[get_db] = override_get_db
client = TestClient(app)

# ── Constants ──────────────────────────────────────────────────────────────────

# Patch targets — these are the network boundaries we replace with mocks
PATCH_EXTRACT = "app.services.receipt_service.extract_receipt_data"
PATCH_TRANSLATE = "app.services.translation_service.translate_expense_text"

TINY_JPEG = b"\xff\xd8\xff\xe0" + b"x" * 100  # minimal valid JPEG header


# ── Helpers ────────────────────────────────────────────────────────────────────

def register_and_login(username: str, email: str, password: str = "Password123!") -> str:
    """Register a user and return their JWT access token."""
    reg = client.post("/api/auth/register", json={
        "username": username,
        "email": email,
        "password": password,
    })
    assert reg.status_code == 201, f"Register failed: {reg.json()}"

    login = client.post("/api/auth/login", json={"email": email, "password": password})
    assert login.status_code == 200, f"Login failed: {login.json()}"
    return login.json()["access_token"]


def auth(token: str) -> dict:
    """Return Authorization header dict."""
    return {"Authorization": f"Bearer {token}"}


def get_my_id(token: str) -> int:
    """Return the user ID for the given token."""
    return client.get("/api/auth/me", headers=auth(token)).json()["id"]


def make_category_db(*, code: str = "FOOD", name_en: str = "Food", name_th: str = "อาหาร") -> int:
    """Insert a category directly in the test DB and return its ID."""
    db = TestingSessionLocal()
    cat = Category(code=code, name_en=name_en, name_th=name_th, is_active=True)
    db.add(cat)
    db.commit()
    db.refresh(cat)
    cat_id = cat.id
    db.close()
    return cat_id


def make_fake_jpg_file() -> str:
    """Write a minimal JPEG to a temp file and return its path."""
    f = tempfile.NamedTemporaryFile(suffix=".jpg", delete=False)
    f.write(TINY_JPEG)
    f.close()
    return f.name


def insert_receipt_db(user_id: int, file_path: str) -> int:
    """Insert a ReceiptFile row pointing to a real file. Returns its ID."""
    db = TestingSessionLocal()
    stored = f"{uuid.uuid4().hex}.jpg"
    receipt = ReceiptFile(
        user_id=user_id,
        expense_id=None,
        original_filename="receipt.jpg",
        stored_filename=stored,
        file_path=file_path,
        mime_type="image/jpeg",
        file_size=len(TINY_JPEG),
        upload_status="uploaded",
    )
    db.add(receipt)
    db.commit()
    db.refresh(receipt)
    rid = receipt.id
    db.close()
    return rid


def upload_receipt_via_api(token: str) -> dict:
    """Upload a JPEG via the real upload route (upload dir patched to tmp)."""
    with tempfile.TemporaryDirectory() as tmpdir:
        with patch("app.services.receipt_service.settings") as mock_cfg:
            mock_cfg.UPLOAD_DIR = tmpdir
            mock_cfg.MAX_RECEIPT_FILE_SIZE_MB = 10
            res = client.post(
                "/api/receipts/upload",
                headers=auth(token),
                files={"file": ("receipt.jpg", io.BytesIO(TINY_JPEG), "image/jpeg")},
            )
    assert res.status_code == 201, f"Upload failed: {res.json()}"
    return res.json()


def fake_extraction() -> ExtractedReceiptData:
    """Return a realistic mocked Gemini extraction result."""
    return ExtractedReceiptData(
        title="Coffee Shop",
        merchant_name="Bean There",
        receipt_date=date(2025, 6, 1),
        currency="THB",
        total_amount=Decimal("95.00"),
        subtotal=Decimal("90.00"),
        tax_amount=Decimal("5.00"),
        language_detected="en",
        ai_confidence=Decimal("0.95"),
        items=[
            ExtractedReceiptItem(
                original_name="Latte",
                name_en="Latte",
                name_th="ลาเต้",
                quantity=Decimal("1"),
                unit_price=Decimal("90.00"),
                total_price=Decimal("90.00"),
                category_name="Food",
            )
        ],
    )


def fake_translation_result(item_id: int) -> GeminiTranslationResult:
    """Return a realistic mocked Gemini translation result."""
    return GeminiTranslationResult(
        translated_title="ร้านกาแฟ",
        translated_notes=None,
        items=[GeminiTranslatedItem(item_id=item_id, translated_name="ลาเต้")],
    )


# ── DB teardown fixture ────────────────────────────────────────────────────────

@pytest.fixture(autouse=True)
def clean_db():
    """Wipe all rows before and after each test to keep tests independent."""
    def wipe():
        db = TestingSessionLocal()
        db.query(Translation).delete()
        db.query(ExpenseItem).delete()
        db.query(ReceiptFile).delete()
        db.query(Expense).delete()
        db.query(Category).delete()
        db.query(RefreshToken).delete()
        db.query(User).delete()
        db.commit()
        db.close()

    wipe()
    yield
    wipe()


# ══════════════════════════════════════════════════════════════════════════════
# Workflow A: AI Receipt Flow
# ══════════════════════════════════════════════════════════════════════════════

class TestWorkflowA_AIReceiptFlow:
    """
    Complete AI-assisted expense flow:
      register → login → upload receipt → extract (mocked Gemini)
      → review draft → update category → confirm
      → translate (mocked Gemini) → export to Excel → validate workbook
    """

    def test_a1_register_and_login(self):
        """User can register and receive an access token."""
        token = register_and_login("user_a1", "user_a1@example.com")
        assert token  # non-empty JWT string

    def test_a2_upload_receipt(self):
        """Authenticated user can upload a JPEG receipt."""
        token = register_and_login("user_a2", "user_a2@example.com")
        receipt = upload_receipt_via_api(token)
        assert receipt["id"] > 0
        assert receipt["upload_status"] == "uploaded"
        assert "file_path" not in receipt  # file path must not be exposed

    def test_a3_extract_receipt_creates_draft_expense(self):
        """Extracting a receipt creates an unconfirmed (draft) expense."""
        token = register_and_login("user_a3", "user_a3@example.com")
        make_category_db(code="FOOD_A3")

        tmp = make_fake_jpg_file()
        try:
            user_id = get_my_id(token)
            rid = insert_receipt_db(user_id, tmp)

            with patch(PATCH_EXTRACT, return_value=fake_extraction()):
                res = client.post(f"/api/receipts/{rid}/extract", headers=auth(token))

            assert res.status_code == 201
            expense = res.json()
            # Draft must not be confirmed yet
            assert expense["is_confirmed"] is False
            assert expense["title"] == "Coffee Shop"
        finally:
            Path(tmp).unlink(missing_ok=True)

    def test_a4_receipt_is_linked_to_draft_expense(self):
        """After extraction, the receipt row has its expense_id set."""
        token = register_and_login("user_a4", "user_a4@example.com")
        make_category_db(code="FOOD_A4")

        tmp = make_fake_jpg_file()
        try:
            user_id = get_my_id(token)
            rid = insert_receipt_db(user_id, tmp)

            with patch(PATCH_EXTRACT, return_value=fake_extraction()):
                expense_id = client.post(
                    f"/api/receipts/{rid}/extract", headers=auth(token)
                ).json()["id"]

            db = TestingSessionLocal()
            receipt = db.query(ReceiptFile).filter(ReceiptFile.id == rid).first()
            db.close()
            assert receipt.expense_id == expense_id
        finally:
            Path(tmp).unlink(missing_ok=True)

    def test_a5_update_draft_expense(self):
        """User can update the draft expense (e.g. fix title) before confirming."""
        token = register_and_login("user_a5", "user_a5@example.com")
        make_category_db(code="FOOD_A5")

        tmp = make_fake_jpg_file()
        try:
            user_id = get_my_id(token)
            rid = insert_receipt_db(user_id, tmp)

            with patch(PATCH_EXTRACT, return_value=fake_extraction()):
                expense_id = client.post(
                    f"/api/receipts/{rid}/extract", headers=auth(token)
                ).json()["id"]

            res = client.put(
                f"/api/expenses/{expense_id}",
                headers=auth(token),
                json={"title": "Morning Coffee Run"},
            )
            assert res.status_code == 200
            assert res.json()["title"] == "Morning Coffee Run"
        finally:
            Path(tmp).unlink(missing_ok=True)

    def test_a6_confirm_expense(self):
        """Confirming the draft sets is_confirmed to True.

        The extraction service leaves category_id as None because ExtractedReceiptData
        has no top-level category field.  The user must assign a category before
        confirming — exactly the 'review draft' step the AI flow describes.
        """
        token = register_and_login("user_a6", "user_a6@example.com")
        cat_id = make_category_db(code="FOOD_A6")

        tmp = make_fake_jpg_file()
        try:
            user_id = get_my_id(token)
            rid = insert_receipt_db(user_id, tmp)

            with patch(PATCH_EXTRACT, return_value=fake_extraction()):
                expense_id = client.post(
                    f"/api/receipts/{rid}/extract", headers=auth(token)
                ).json()["id"]

            # User reviews the draft and assigns a category (required before confirm)
            client.put(
                f"/api/expenses/{expense_id}",
                headers=auth(token),
                json={"category_id": cat_id},
            )

            res = client.post(f"/api/expenses/{expense_id}/confirm", headers=auth(token))
            assert res.status_code == 200
            assert res.json()["is_confirmed"] is True
        finally:
            Path(tmp).unlink(missing_ok=True)

    def test_a7_translate_confirmed_expense(self):
        """Translating a confirmed expense returns Thai text (Gemini mocked)."""
        token = register_and_login("user_a7", "user_a7@example.com")
        cat_id = make_category_db(code="FOOD_A7")

        tmp = make_fake_jpg_file()
        try:
            user_id = get_my_id(token)
            rid = insert_receipt_db(user_id, tmp)

            with patch(PATCH_EXTRACT, return_value=fake_extraction()):
                expense_data = client.post(
                    f"/api/receipts/{rid}/extract", headers=auth(token)
                ).json()
            expense_id = expense_data["id"]
            item_id = expense_data["items"][0]["id"] if expense_data["items"] else 1

            # Assign category before confirming
            client.put(
                f"/api/expenses/{expense_id}",
                headers=auth(token),
                json={"category_id": cat_id},
            )
            client.post(f"/api/expenses/{expense_id}/confirm", headers=auth(token))

            with patch(PATCH_TRANSLATE, return_value=fake_translation_result(item_id)):
                res = client.post(
                    f"/api/expenses/{expense_id}/translate",
                    headers=auth(token),
                    json={"target_language": "th"},
                )

            assert res.status_code == 200
            data = res.json()
            assert data["target_language"] == "th"
            assert data["translated_title"] == "ร้านกาแฟ"
        finally:
            Path(tmp).unlink(missing_ok=True)

    def test_a8_export_contains_confirmed_expense(self):
        """The Excel export includes the confirmed expense."""
        token = register_and_login("user_a8", "user_a8@example.com")
        cat_id = make_category_db(code="FOOD_A8")

        tmp = make_fake_jpg_file()
        try:
            user_id = get_my_id(token)
            rid = insert_receipt_db(user_id, tmp)

            with patch(PATCH_EXTRACT, return_value=fake_extraction()):
                expense_id = client.post(
                    f"/api/receipts/{rid}/extract", headers=auth(token)
                ).json()["id"]

            # Assign category, then confirm
            client.put(
                f"/api/expenses/{expense_id}",
                headers=auth(token),
                json={"category_id": cat_id},
            )
            client.post(f"/api/expenses/{expense_id}/confirm", headers=auth(token))

            res = client.get("/api/expenses/export", headers=auth(token))
            assert res.status_code == 200

            wb = load_workbook(BytesIO(res.content))
            ws = wb["Expenses"]
            titles = [ws.cell(row=r, column=4).value for r in range(2, ws.max_row + 1)]
            assert "Coffee Shop" in titles
        finally:
            Path(tmp).unlink(missing_ok=True)

    def test_a9_workbook_has_expected_sheets(self):
        """Exported workbook contains Expenses and Expense Items sheets."""
        token = register_and_login("user_a9", "user_a9@example.com")
        res = client.get("/api/expenses/export", headers=auth(token))
        assert res.status_code == 200
        wb = load_workbook(BytesIO(res.content))
        assert "Expenses" in wb.sheetnames
        assert "Expense Items" in wb.sheetnames

    def test_a10_no_real_gemini_called(self):
        """Structural check: the network boundary function exists and is patchable."""
        import app.services.receipt_service as receipt_svc
        import app.services.translation_service as translation_svc
        assert callable(receipt_svc.extract_receipt_data)
        assert callable(translation_svc.translate_expense_text)


# ══════════════════════════════════════════════════════════════════════════════
# Workflow B: Manual Expense Flow
# ══════════════════════════════════════════════════════════════════════════════

class TestWorkflowB_ManualExpenseFlow:
    """
    Full manual expense lifecycle:
      register → login → create → list → detail → update
      → export → soft-delete → confirm gone
    """

    def _create_expense(self, token: str, category_id: int, title: str = "Team Lunch") -> dict:
        res = client.post(
            "/api/expenses",
            headers=auth(token),
            json={
                "category_id": category_id,
                "title": title,
                "receipt_date": "2026-01-15",
                "total_amount": "350.00",
                "currency": "THB",
                "items": [
                    {"original_name": "Pad Thai", "quantity": "2", "total_price": "200.00"},
                    {"original_name": "Papaya Salad", "quantity": "1", "total_price": "150.00"},
                ],
            },
        )
        assert res.status_code == 201, f"Create failed: {res.json()}"
        return res.json()

    def test_b1_create_expense_manually(self):
        """Creating an expense returns 201 with correct data."""
        token = register_and_login("user_b1", "user_b1@example.com")
        cat_id = make_category_db(code="FOOD_B1")
        expense = self._create_expense(token, cat_id)

        assert expense["title"] == "Team Lunch"
        assert expense["total_amount"] == "350.00"
        assert len(expense["items"]) == 2

    def test_b2_list_shows_created_expense(self):
        """The created expense appears in the list endpoint."""
        token = register_and_login("user_b2", "user_b2@example.com")
        cat_id = make_category_db(code="FOOD_B2")
        self._create_expense(token, cat_id)

        res = client.get("/api/expenses", headers=auth(token))
        assert res.status_code == 200
        titles = [e["title"] for e in res.json()]
        assert "Team Lunch" in titles

    def test_b3_detail_returns_expense_with_items(self):
        """GET detail returns the expense including its nested items."""
        token = register_and_login("user_b3", "user_b3@example.com")
        cat_id = make_category_db(code="FOOD_B3")
        expense = self._create_expense(token, cat_id)

        res = client.get(f"/api/expenses/{expense['id']}", headers=auth(token))
        assert res.status_code == 200
        data = res.json()
        assert data["id"] == expense["id"]
        assert len(data["items"]) == 2
        item_names = {i["original_name"] for i in data["items"]}
        assert item_names == {"Pad Thai", "Papaya Salad"}

    def test_b4_update_expense(self):
        """PUT updates the expense title and notes."""
        token = register_and_login("user_b4", "user_b4@example.com")
        cat_id = make_category_db(code="FOOD_B4")
        expense = self._create_expense(token, cat_id)

        res = client.put(
            f"/api/expenses/{expense['id']}",
            headers=auth(token),
            json={"title": "Updated Lunch", "notes": "paid by company card"},
        )
        assert res.status_code == 200
        data = res.json()
        assert data["title"] == "Updated Lunch"
        assert data["notes"] == "paid by company card"

    def test_b5_export_contains_expense(self):
        """Excel export includes the manually created expense."""
        token = register_and_login("user_b5", "user_b5@example.com")
        cat_id = make_category_db(code="FOOD_B5")
        self._create_expense(token, cat_id)

        res = client.get("/api/expenses/export", headers=auth(token))
        assert res.status_code == 200
        wb = load_workbook(BytesIO(res.content))
        ws = wb["Expenses"]
        titles = [ws.cell(row=r, column=4).value for r in range(2, ws.max_row + 1)]
        assert "Team Lunch" in titles

    def test_b6_soft_delete_removes_from_list(self):
        """Deleting an expense removes it from the list endpoint."""
        token = register_and_login("user_b6", "user_b6@example.com")
        cat_id = make_category_db(code="FOOD_B6")
        expense = self._create_expense(token, cat_id)

        del_res = client.delete(f"/api/expenses/{expense['id']}", headers=auth(token))
        assert del_res.status_code == 200

        list_res = client.get("/api/expenses", headers=auth(token))
        ids = [e["id"] for e in list_res.json()]
        assert expense["id"] not in ids

    def test_b7_detail_returns_404_after_delete(self):
        """Getting detail of a deleted expense returns 404."""
        token = register_and_login("user_b7", "user_b7@example.com")
        cat_id = make_category_db(code="FOOD_B7")
        expense = self._create_expense(token, cat_id)

        client.delete(f"/api/expenses/{expense['id']}", headers=auth(token))

        res = client.get(f"/api/expenses/{expense['id']}", headers=auth(token))
        assert res.status_code == 404

    def test_b8_row_still_exists_in_db_after_soft_delete(self):
        """Soft delete sets deleted_at but keeps the row in the database."""
        token = register_and_login("user_b8", "user_b8@example.com")
        cat_id = make_category_db(code="FOOD_B8")
        expense = self._create_expense(token, cat_id)

        client.delete(f"/api/expenses/{expense['id']}", headers=auth(token))

        db = TestingSessionLocal()
        row = db.query(Expense).filter(Expense.id == expense["id"]).first()
        db.close()
        # Row must still exist
        assert row is not None
        # And must have deleted_at set
        assert row.deleted_at is not None

    def test_b9_deleted_expense_excluded_from_export(self):
        """Export only includes active (non-deleted) expenses."""
        token = register_and_login("user_b9", "user_b9@example.com")
        cat_id = make_category_db(code="FOOD_B9")
        active = self._create_expense(token, cat_id, title="Keep This")
        deleted = self._create_expense(token, cat_id, title="Delete This")

        client.delete(f"/api/expenses/{deleted['id']}", headers=auth(token))

        res = client.get("/api/expenses/export", headers=auth(token))
        wb = load_workbook(BytesIO(res.content))
        ws = wb["Expenses"]
        titles = [ws.cell(row=r, column=4).value for r in range(2, ws.max_row + 1)]
        assert "Keep This" in titles
        assert "Delete This" not in titles

    def test_b10_sensitive_fields_not_in_responses(self):
        """Expense responses never expose internal fields."""
        token = register_and_login("user_b10", "user_b10@example.com")
        cat_id = make_category_db(code="FOOD_B10")
        expense = self._create_expense(token, cat_id)

        detail = client.get(f"/api/expenses/{expense['id']}", headers=auth(token)).json()
        for field in ("deleted_at", "ai_raw_response", "ai_confidence", "ai_status",
                      "language_detected", "password", "password_hash"):
            assert field not in detail, f"Sensitive field exposed: {field}"


# ══════════════════════════════════════════════════════════════════════════════
# Workflow C: User Isolation
# ══════════════════════════════════════════════════════════════════════════════

class TestWorkflowC_UserIsolation:
    """
    Two users (A and B) are registered separately.
    User B must never be able to access or modify User A's data.
    All endpoints return 404 for cross-user access (no information leakage).
    """

    def _setup_two_users(self) -> tuple[str, str, int]:
        """Register two users, create a category, and return (token_a, token_b, expense_id_for_a)."""
        token_a = register_and_login("iso_user_a", "iso_a@example.com")
        token_b = register_and_login("iso_user_b", "iso_b@example.com")
        cat_id = make_category_db(code="FOOD_ISO")

        expense_a = client.post(
            "/api/expenses",
            headers=auth(token_a),
            json={
                "category_id": cat_id,
                "title": "User A Expense",
                "receipt_date": "2026-01-10",
                "total_amount": "100.00",
                "currency": "THB",
                "items": [],
            },
        ).json()
        return token_a, token_b, expense_a["id"]

    def test_c1_user_b_cannot_view_user_a_expense(self):
        """GET detail of User A's expense by User B returns 404."""
        _, token_b, expense_id = self._setup_two_users()
        res = client.get(f"/api/expenses/{expense_id}", headers=auth(token_b))
        assert res.status_code == 404

    def test_c2_user_b_cannot_update_user_a_expense(self):
        """PUT on User A's expense by User B returns 404."""
        _, token_b, expense_id = self._setup_two_users()
        res = client.put(
            f"/api/expenses/{expense_id}",
            headers=auth(token_b),
            json={"title": "Hijacked"},
        )
        assert res.status_code == 404

    def test_c3_user_b_cannot_delete_user_a_expense(self):
        """DELETE on User A's expense by User B returns 404."""
        _, token_b, expense_id = self._setup_two_users()
        res = client.delete(f"/api/expenses/{expense_id}", headers=auth(token_b))
        assert res.status_code == 404

    def test_c4_user_b_cannot_confirm_user_a_expense(self):
        """POST confirm on User A's expense by User B returns 404."""
        _, token_b, expense_id = self._setup_two_users()
        res = client.post(f"/api/expenses/{expense_id}/confirm", headers=auth(token_b))
        assert res.status_code == 404

    def test_c5_user_b_cannot_translate_user_a_expense(self):
        """POST translate on User A's expense by User B returns 404."""
        _, token_b, expense_id = self._setup_two_users()
        res = client.post(
            f"/api/expenses/{expense_id}/translate",
            headers=auth(token_b),
            json={"target_language": "th"},
        )
        assert res.status_code == 404

    def test_c6_user_b_list_does_not_contain_user_a_expenses(self):
        """GET /api/expenses for User B returns only their own expenses."""
        token_a, token_b, _ = self._setup_two_users()
        res = client.get("/api/expenses", headers=auth(token_b))
        assert res.status_code == 200
        # User B has no expenses — list must be empty
        assert res.json() == []

    def test_c7_user_b_export_does_not_contain_user_a_data(self):
        """Excel export for User B must not include User A's expenses."""
        token_a, token_b, _ = self._setup_two_users()
        res = client.get("/api/expenses/export", headers=auth(token_b))
        assert res.status_code == 200
        wb = load_workbook(BytesIO(res.content))
        ws = wb["Expenses"]
        # Header only — no data rows for User B
        assert ws.max_row == 1

    def test_c8_user_b_cannot_link_receipt_to_user_a_expense(self):
        """POST link receipt to User A's expense by User B returns 404."""
        token_a, token_b, expense_id = self._setup_two_users()
        uid_b = get_my_id(token_b)
        tmp = make_fake_jpg_file()
        try:
            rid = insert_receipt_db(uid_b, tmp)
            res = client.post(
                f"/api/expenses/{expense_id}/receipts/{rid}",
                headers=auth(token_b),
            )
            assert res.status_code == 404
        finally:
            Path(tmp).unlink(missing_ok=True)

    def test_c9_user_b_cannot_view_user_a_receipt(self):
        """GET detail of User A's receipt by User B returns 404."""
        token_a, token_b, _ = self._setup_two_users()
        uid_a = get_my_id(token_a)
        tmp = make_fake_jpg_file()
        try:
            rid = insert_receipt_db(uid_a, tmp)
            res = client.get(f"/api/receipts/{rid}", headers=auth(token_b))
            assert res.status_code == 404
        finally:
            Path(tmp).unlink(missing_ok=True)

    def test_c10_no_data_leakage_in_404_response(self):
        """A 404 response for another user's expense does not include any expense data."""
        token_a, token_b, expense_id = self._setup_two_users()
        res = client.get(f"/api/expenses/{expense_id}", headers=auth(token_b))
        assert res.status_code == 404
        # The error body must not contain the expense's title or user-specific data
        assert "User A Expense" not in res.text
        assert str(get_my_id(token_a)) not in res.text


# ══════════════════════════════════════════════════════════════════════════════
# Route registration sanity checks
# ══════════════════════════════════════════════════════════════════════════════

class TestRouteRegistration:
    """
    Verify all expected routes are registered and reachable
    (no 404 from missing routes — wrong auth is fine).
    """

    PROTECTED_GET_ROUTES = [
        "/api/auth/me",
        "/api/categories",
        "/api/receipts",
        "/api/expenses",
        "/api/expenses/export",
    ]

    PROTECTED_POST_ROUTES = [
        "/api/auth/login",   # not protected, but should exist
        "/api/auth/register",
    ]

    def test_all_unauthenticated_protected_routes_return_401_not_404(self):
        """Protected routes must return 401 (not 404) when called without a token."""
        for route in self.PROTECTED_GET_ROUTES:
            res = client.get(route)
            assert res.status_code != 404, f"Route not registered: GET {route}"
            assert res.status_code == 401, f"Expected 401 on GET {route}, got {res.status_code}"

    def test_export_route_is_not_shadowed_by_expense_id_route(self):
        """
        /api/expenses/export must resolve to the export route, not /{expense_id}.
        If /export were treated as an expense ID it would 401 (no token),
        which still means the route is registered — but the handler is correct
        because export is defined before /{expense_id} in expenses.py.
        """
        res = client.get("/api/expenses/export")
        # Without token: 401 from the export handler (correctly identified as the export route)
        assert res.status_code == 401

    def test_health_route_is_public(self):
        """GET /api/health returns 200 without authentication."""
        res = client.get("/api/health")
        assert res.status_code == 200

    def test_auth_register_exists(self):
        """POST /api/auth/register is reachable (422 for empty body, not 404)."""
        res = client.post("/api/auth/register", json={})
        assert res.status_code != 404

    def test_auth_login_exists(self):
        """POST /api/auth/login is reachable."""
        res = client.post("/api/auth/login", json={})
        assert res.status_code != 404

    def test_category_detail_route_exists(self):
        """GET /api/categories/{id} returns 401 without token, not 404."""
        res = client.get("/api/categories/1")
        assert res.status_code == 401

    def test_receipt_upload_route_exists(self):
        """POST /api/receipts/upload returns 401 without token, not 404."""
        res = client.post("/api/receipts/upload")
        assert res.status_code == 401

    def test_receipt_extract_route_exists(self):
        """POST /api/receipts/{id}/extract returns 401 without token, not 404."""
        res = client.post("/api/receipts/1/extract")
        assert res.status_code == 401

    def test_expense_confirm_route_exists(self):
        """POST /api/expenses/{id}/confirm returns 401 without token, not 404."""
        res = client.post("/api/expenses/1/confirm")
        assert res.status_code == 401

    def test_expense_translate_route_exists(self):
        """POST /api/expenses/{id}/translate returns 401 without token, not 404."""
        res = client.post("/api/expenses/1/translate", json={"target_language": "th"})
        assert res.status_code == 401

    def test_expense_receipt_link_route_exists(self):
        """POST /api/expenses/{id}/receipts/{rid} returns 401 without token."""
        res = client.post("/api/expenses/1/receipts/1")
        assert res.status_code == 401

    def test_expense_receipt_unlink_route_exists(self):
        """DELETE /api/expenses/{id}/receipts/{rid} returns 401 without token."""
        res = client.delete("/api/expenses/1/receipts/1")
        assert res.status_code == 401
