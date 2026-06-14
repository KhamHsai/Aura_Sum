"""Service layer for expense management."""

from datetime import datetime, timezone
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from app.models.category import Category
from app.models.expense import Expense
from app.models.expense_item import ExpenseItem
from app.models.receipt_file import ReceiptFile
from app.schemas.expense import ExpenseCreate, ExpenseResponse, ExpenseUpdate
from app.schemas.expense_item import ExpenseItemResponse


class ExpenseServiceError(Exception):
    """Raised when business validation fails in the expense service."""

    def __init__(self, message: str, status_code: int = 400):
        self.message = message
        self.status_code = status_code
        super().__init__(message)


def _get_valid_category(db: Session, category_id: int, error_message: str) -> Category:
    """Return a category that exists, is active, and is not soft-deleted."""
    category = (
        db.query(Category)
        .filter(
            Category.id == category_id,
            Category.is_active == True,
            Category.deleted_at.is_(None),
        )
        .first()
    )
    if not category:
        raise ExpenseServiceError(error_message, status_code=404)
    return category


def create_expense(db: Session, user_id: int, data: ExpenseCreate) -> ExpenseResponse:
    """
    Create one expense and its optional items for the authenticated user.

    Validates categories, writes everything in one transaction,
    and rolls back fully if anything fails.
    Returns an ExpenseResponse built while the session is still open.
    """
    # 1. Validate the main expense category
    _get_valid_category(db, data.category_id, "Category not found")

    # 2. Validate each item category before touching the database
    for item_data in data.items:
        if item_data.category_id is not None:
            _get_valid_category(db, item_data.category_id, "Item category not found")

    try:
        # 3. Create the Expense — user_id always comes from the auth layer
        expense = Expense(
            user_id=user_id,
            category_id=data.category_id,
            title=data.title,
            merchant_name=data.merchant_name,
            receipt_number=data.receipt_number,
            receipt_date=data.receipt_date,
            payment_method=data.payment_method,
            currency=data.currency,
            subtotal=data.subtotal,
            tax_amount=data.tax_amount,
            discount_amount=data.discount_amount,
            total_amount=data.total_amount,
            notes=data.notes,
            input_method="manual",  # only supported value for user-created expenses
        )

        # 4. Flush to get expense.id without committing yet
        db.add(expense)
        db.flush()

        # 5. Create each ExpenseItem using the flushed expense.id, collecting them
        created_items: list[ExpenseItem] = []
        for item_data in data.items:
            item = ExpenseItem(
                expense_id=expense.id,
                category_id=item_data.category_id,
                original_name=item_data.original_name,
                name_en=item_data.name_en,
                name_th=item_data.name_th,
                quantity=item_data.quantity,
                unit=item_data.unit,
                unit_price=item_data.unit_price,
                discount_amount=item_data.discount_amount,
                total_price=item_data.total_price,
            )
            db.add(item)
            created_items.append(item)

        # 6. Flush items to get their IDs, then commit once
        db.flush()
        db.commit()

        # 7. Refresh expense and all items to load server-generated fields
        #    (created_at, updated_at) while the session is still open.
        db.refresh(expense)
        for item in created_items:
            db.refresh(item)

        # 8. Build the Pydantic response from the in-memory objects while the
        #    session is still open — avoids lazy-load issues after session closes.
        item_responses = [ExpenseItemResponse.model_validate(item) for item in created_items]
        response = ExpenseResponse(
            id=expense.id,
            user_id=expense.user_id,
            category_id=expense.category_id,
            title=expense.title,
            merchant_name=expense.merchant_name,
            receipt_number=expense.receipt_number,
            receipt_date=expense.receipt_date,
            payment_method=expense.payment_method,
            currency=expense.currency,
            subtotal=expense.subtotal,
            tax_amount=expense.tax_amount,
            discount_amount=expense.discount_amount,
            total_amount=expense.total_amount,
            notes=expense.notes,
            is_confirmed=expense.is_confirmed,
            created_at=expense.created_at,
            updated_at=expense.updated_at,
            items=item_responses,
        )
        return response

    except ExpenseServiceError:
        db.rollback()
        raise
    except Exception:
        db.rollback()
        raise


# ── Helpers ────────────────────────────────────────────────────────────────────

def _get_non_deleted_items(db: Session, expense_id: int) -> list[ExpenseItem]:
    """Return non-deleted items for an expense, ordered by id ascending."""
    return (
        db.query(ExpenseItem)
        .filter(
            ExpenseItem.expense_id == expense_id,
            ExpenseItem.deleted_at.is_(None),
        )
        .order_by(ExpenseItem.id.asc())
        .all()
    )


def _build_expense_response(expense: Expense, items: list[ExpenseItem]) -> ExpenseResponse:
    """Build an ExpenseResponse from ORM objects while the session is open."""
    item_responses = [ExpenseItemResponse.model_validate(item) for item in items]
    return ExpenseResponse(
        id=expense.id,
        user_id=expense.user_id,
        category_id=expense.category_id,
        title=expense.title,
        merchant_name=expense.merchant_name,
        receipt_number=expense.receipt_number,
        receipt_date=expense.receipt_date,
        payment_method=expense.payment_method,
        currency=expense.currency,
        subtotal=expense.subtotal,
        tax_amount=expense.tax_amount,
        discount_amount=expense.discount_amount,
        total_amount=expense.total_amount,
        notes=expense.notes,
        is_confirmed=expense.is_confirmed,
        created_at=expense.created_at,
        updated_at=expense.updated_at,
        items=item_responses,
    )


# ── Read functions ─────────────────────────────────────────────────────────────

def get_user_expenses(db: Session, user_id: int) -> list[ExpenseResponse]:
    """
    Return all non-deleted expenses belonging to one user, newest first.

    Each expense includes only its non-deleted items.
    Returns an empty list when the user has no expenses.
    """
    expenses = (
        db.query(Expense)
        .filter(
            Expense.user_id == user_id,
            Expense.deleted_at.is_(None),
        )
        .order_by(Expense.created_at.desc(), Expense.id.desc())
        .all()
    )

    result = []
    for expense in expenses:
        items = _get_non_deleted_items(db, expense.id)
        result.append(_build_expense_response(expense, items))
    return result


def get_user_expense_by_id(
    db: Session,
    user_id: int,
    expense_id: int,
) -> ExpenseResponse | None:
    """
    Return one owned, non-deleted expense with its non-deleted items.

    Returns None when the expense does not exist, belongs to another user,
    or has been soft-deleted — so the caller can safely return 404 without
    revealing whether another user's expense exists.
    """
    expense = (
        db.query(Expense)
        .filter(
            Expense.id == expense_id,
            Expense.user_id == user_id,
            Expense.deleted_at.is_(None),
        )
        .first()
    )

    if expense is None:
        return None

    items = _get_non_deleted_items(db, expense.id)
    return _build_expense_response(expense, items)


# ── Update ─────────────────────────────────────────────────────────────────────

def update_user_expense(
    db: Session,
    user_id: int,
    expense_id: int,
    data: ExpenseUpdate,
) -> ExpenseResponse | None:
    """
    Partially update one owned, non-deleted expense.

    Only fields present in the request body are written.
    Items are replaced only when the 'items' key was provided.
    Returns None when the expense is unavailable; the route converts this to 404.
    """
    expense = (
        db.query(Expense)
        .filter(
            Expense.id == expense_id,
            Expense.user_id == user_id,
            Expense.deleted_at.is_(None),
        )
        .first()
    )
    if expense is None:
        return None

    updates = data.model_dump(exclude_unset=True)

    # Validate new category before touching anything
    if "category_id" in updates:
        _get_valid_category(db, updates["category_id"], "Category not found")

    # Validate all new item categories before touching anything
    new_items_data = updates.pop("items", None)  # None means not provided
    if new_items_data is not None:
        for item_dict in new_items_data:
            cat_id = item_dict.get("category_id")
            if cat_id is not None:
                _get_valid_category(db, cat_id, "Item category not found")

    # Simple-field editable names that the client is allowed to change
    _EDITABLE = {
        "category_id", "title", "merchant_name", "receipt_number", "receipt_date",
        "payment_method", "currency", "subtotal", "tax_amount", "discount_amount",
        "total_amount", "notes",
    }

    try:
        # Apply simple field updates
        for field, value in updates.items():
            if field in _EDITABLE:
                setattr(expense, field, value)

        # Replace items only when the key was present in the request
        if new_items_data is not None:
            now = datetime.now(timezone.utc).replace(tzinfo=None)

            # Soft-delete all existing active items
            active_items = _get_non_deleted_items(db, expense.id)
            for item in active_items:
                item.deleted_at = now

            # Create all new item records
            created_items: list[ExpenseItem] = []
            for item_dict in new_items_data:
                item = ExpenseItem(
                    expense_id=expense.id,
                    category_id=item_dict.get("category_id"),
                    original_name=item_dict["original_name"],
                    name_en=item_dict.get("name_en"),
                    name_th=item_dict.get("name_th"),
                    quantity=item_dict["quantity"],
                    unit=item_dict.get("unit"),
                    unit_price=item_dict.get("unit_price"),
                    discount_amount=item_dict.get("discount_amount", 0),
                    total_price=item_dict["total_price"],
                )
                db.add(item)
                created_items.append(item)

            db.flush()
            for item in created_items:
                db.refresh(item)

        db.flush()
        db.commit()
        db.refresh(expense)

        items = _get_non_deleted_items(db, expense.id)
        return _build_expense_response(expense, items)

    except ExpenseServiceError:
        db.rollback()
        raise
    except Exception:
        db.rollback()
        raise


# ── Delete ─────────────────────────────────────────────────────────────────────

def delete_user_expense(
    db: Session,
    user_id: int,
    expense_id: int,
) -> bool:
    """
    Soft-delete an owned expense and all its active items.

    Returns False when the expense is unavailable (missing, wrong owner, already deleted).
    Does not permanently delete rows or touch receipt files.
    """
    expense = (
        db.query(Expense)
        .filter(
            Expense.id == expense_id,
            Expense.user_id == user_id,
            Expense.deleted_at.is_(None),
        )
        .first()
    )
    if expense is None:
        return False

    try:
        now = datetime.now(timezone.utc).replace(tzinfo=None)

        expense.deleted_at = now

        active_items = _get_non_deleted_items(db, expense.id)
        for item in active_items:
            item.deleted_at = now

        db.commit()
        return True

    except Exception:
        db.rollback()
        raise


# ── Receipt linking ────────────────────────────────────────────────────────────

def _get_owned_expense(db: Session, user_id: int, expense_id: int) -> Expense:
    """Return the owned, non-deleted expense or raise 404 ExpenseServiceError."""
    expense = (
        db.query(Expense)
        .filter(
            Expense.id == expense_id,
            Expense.user_id == user_id,
            Expense.deleted_at.is_(None),
        )
        .first()
    )
    if expense is None:
        raise ExpenseServiceError("Expense or receipt not found", status_code=404)
    return expense


def _get_owned_receipt(db: Session, user_id: int, receipt_id: int) -> ReceiptFile:
    """Return the owned, non-deleted receipt or raise 404 ExpenseServiceError."""
    receipt = (
        db.query(ReceiptFile)
        .filter(
            ReceiptFile.id == receipt_id,
            ReceiptFile.user_id == user_id,
            ReceiptFile.deleted_at.is_(None),
        )
        .first()
    )
    if receipt is None:
        raise ExpenseServiceError("Expense or receipt not found", status_code=404)
    return receipt


def link_receipt_to_expense(
    db: Session,
    user_id: int,
    expense_id: int,
    receipt_id: int,
) -> ReceiptFile:
    """
    Link an owned receipt to an owned expense.

    - If already linked to the same expense, returns success (idempotent).
    - If linked to a different expense, raises 409.
    - Otherwise sets receipt.expense_id and commits once.
    """
    expense = _get_owned_expense(db, user_id, expense_id)
    receipt = _get_owned_receipt(db, user_id, receipt_id)

    # Already linked to the same expense — nothing to do
    if receipt.expense_id == expense.id:
        return receipt

    # Linked to a different expense — require explicit unlink first
    if receipt.expense_id is not None:
        raise ExpenseServiceError(
            "Receipt is already linked to another expense", status_code=409
        )

    try:
        receipt.expense_id = expense.id
        db.commit()
        db.refresh(receipt)
        return receipt
    except Exception:
        db.rollback()
        raise


def unlink_receipt_from_expense(
    db: Session,
    user_id: int,
    expense_id: int,
    receipt_id: int,
) -> ReceiptFile:
    """
    Remove the link between an owned receipt and an owned expense.

    - If the receipt is not linked to this expense, raises 409.
    - Otherwise sets receipt.expense_id to None and commits once.
    """
    expense = _get_owned_expense(db, user_id, expense_id)
    receipt = _get_owned_receipt(db, user_id, receipt_id)

    # Receipt must be linked to this specific expense
    if receipt.expense_id != expense.id:
        raise ExpenseServiceError(
            "Receipt is not linked to this expense", status_code=409
        )

    try:
        receipt.expense_id = None
        db.commit()
        db.refresh(receipt)
        return receipt
    except Exception:
        db.rollback()
        raise


# ── Confirm ────────────────────────────────────────────────────────────────────

# ── Excel Export ───────────────────────────────────────────────────────────────

def _category_name(category) -> str:
    """Return a readable category name, falling back to 'Uncategorized'."""
    if category is None:
        return "Uncategorized"
    return category.name_en or category.name_th or "Uncategorized"


def _apply_header_style(ws) -> None:
    """Bold the header row, freeze it, and enable auto-filter."""
    bold = Font(bold=True)
    for cell in ws[1]:
        cell.font = bold
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def _set_column_widths(ws, widths: list[int]) -> None:
    """Set column widths by position (1-indexed list)."""
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width


def export_user_expenses_to_excel(db: Session, user_id: int) -> BytesIO:
    """
    Build an in-memory Excel workbook with two worksheets:
      - Expenses  : one row per active expense
      - Expense Items : one row per active item across all active expenses

    Only the authenticated user's non-deleted expenses and items are included.
    Returns a BytesIO stream ready to be sent as a file download.
    """
    # ── 1. Query data ──────────────────────────────────────────────────────────
    expenses = (
        db.query(Expense)
        .filter(
            Expense.user_id == user_id,
            Expense.deleted_at.is_(None),
        )
        .order_by(Expense.created_at.asc(), Expense.id.asc())
        .all()
    )

    expense_ids = [e.id for e in expenses]

    items = []
    if expense_ids:
        items = (
            db.query(ExpenseItem)
            .filter(
                ExpenseItem.expense_id.in_(expense_ids),
                ExpenseItem.deleted_at.is_(None),
            )
            .order_by(ExpenseItem.expense_id.asc(), ExpenseItem.id.asc())
            .all()
        )

    # ── 2. Build workbook ──────────────────────────────────────────────────────
    wb = Workbook()

    # Remove the default empty sheet
    default_sheet = wb.active
    wb.remove(default_sheet)

    # ── 3. Expenses worksheet ──────────────────────────────────────────────────
    ws_exp = wb.create_sheet("Expenses")

    expense_headers = [
        "Expense ID", "Receipt Date", "Receipt Time", "Title", "Merchant Name",
        "Category", "Receipt Number", "Document Type", "Payment Method", "Currency",
        "Subtotal", "Tax Amount", "Discount Amount", "Total Amount",
        "Input Method", "Language", "Confirmed", "Notes", "Created At", "Updated At",
    ]
    ws_exp.append(expense_headers)

    for exp in expenses:
        ws_exp.append([
            exp.id,
            exp.receipt_date,            # date → formatted below
            exp.receipt_time,            # time → formatted below
            exp.title,
            exp.merchant_name,
            _category_name(exp.category),
            exp.receipt_number,
            exp.document_type,
            exp.payment_method,
            exp.currency,
            float(exp.subtotal) if exp.subtotal is not None else None,
            float(exp.tax_amount) if exp.tax_amount is not None else None,
            float(exp.discount_amount) if exp.discount_amount is not None else None,
            float(exp.total_amount) if exp.total_amount is not None else None,
            exp.input_method,
            exp.language_detected,
            "Yes" if exp.is_confirmed else "No",
            exp.notes,
            exp.created_at,
            exp.updated_at,
        ])

    # Apply number/date formats to money and date columns
    MONEY_COLS_EXP = [11, 12, 13, 14]   # Subtotal … Total Amount (1-indexed)
    DATE_COL_EXP = 2                     # Receipt Date
    TIME_COL_EXP = 3                     # Receipt Time
    DT_COLS_EXP = [19, 20]              # Created At, Updated At

    for row in ws_exp.iter_rows(min_row=2):
        for col_idx in MONEY_COLS_EXP:
            cell = row[col_idx - 1]
            if cell.value is not None:
                cell.number_format = "0.00"
        date_cell = row[DATE_COL_EXP - 1]
        if date_cell.value is not None:
            date_cell.number_format = "yyyy-mm-dd"
        time_cell = row[TIME_COL_EXP - 1]
        if time_cell.value is not None:
            time_cell.number_format = "hh:mm:ss"
        for col_idx in DT_COLS_EXP:
            cell = row[col_idx - 1]
            if cell.value is not None:
                cell.number_format = "yyyy-mm-dd hh:mm:ss"

    _apply_header_style(ws_exp)
    _set_column_widths(ws_exp, [
        10, 12, 10, 28, 22,   # ID … Merchant Name
        16, 14, 16, 16, 8,    # Category … Currency
        12, 12, 14, 14,       # Subtotal … Total Amount
        14, 10, 10, 30,       # Input Method … Notes
        18, 18,               # Created At, Updated At
    ])

    # ── 4. Expense Items worksheet ─────────────────────────────────────────────
    ws_items = wb.create_sheet("Expense Items")

    item_headers = [
        "Expense ID", "Item ID", "Original Name", "English Name", "Thai Name",
        "Category", "Quantity", "Unit", "Unit Price", "Discount Amount",
        "Total Price", "Created At", "Updated At",
    ]
    ws_items.append(item_headers)

    for item in items:
        ws_items.append([
            item.expense_id,
            item.id,
            item.original_name,
            item.name_en,
            item.name_th,
            _category_name(item.category),
            float(item.quantity) if item.quantity is not None else None,
            item.unit,
            float(item.unit_price) if item.unit_price is not None else None,
            float(item.discount_amount) if item.discount_amount is not None else None,
            float(item.total_price) if item.total_price is not None else None,
            item.created_at,
            item.updated_at,
        ])

    MONEY_COLS_ITEMS = [9, 10, 11]   # Unit Price … Total Price
    QTY_COL = 7
    DT_COLS_ITEMS = [12, 13]

    for row in ws_items.iter_rows(min_row=2):
        for col_idx in MONEY_COLS_ITEMS:
            cell = row[col_idx - 1]
            if cell.value is not None:
                cell.number_format = "0.00"
        qty_cell = row[QTY_COL - 1]
        if qty_cell.value is not None:
            qty_cell.number_format = "0.000"
        for col_idx in DT_COLS_ITEMS:
            cell = row[col_idx - 1]
            if cell.value is not None:
                cell.number_format = "yyyy-mm-dd hh:mm:ss"

    _apply_header_style(ws_items)
    _set_column_widths(ws_items, [
        10, 8, 28, 22, 22,    # Expense ID … Thai Name
        16, 10, 10, 12, 14,   # Category … Discount Amount
        12, 18, 18,           # Total Price … Updated At
    ])

    # ── 5. Save to memory and return ───────────────────────────────────────────
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream


def confirm_user_expense(
    db: Session,
    user_id: int,
    expense_id: int,
) -> ExpenseResponse:
    """
    Confirm an AI-extracted draft expense after the user has reviewed it.

    Checks (in order):
    1. Expense exists, belongs to user, and is not soft-deleted (404).
    2. Expense was created by AI extraction (409).
    3. Expense is not already confirmed (409).
    4. category_id is set (422).
    5. category exists, is active, and is not soft-deleted (422).
    6. title is present and not blank (422).
    7. total_amount is present and zero or greater (422).

    On success, sets is_confirmed = True and commits once.
    Rolls back if the commit fails.
    """
    # 1. Ownership check
    expense = (
        db.query(Expense)
        .filter(
            Expense.id == expense_id,
            Expense.user_id == user_id,
            Expense.deleted_at.is_(None),
        )
        .first()
    )
    if expense is None:
        raise ExpenseServiceError("Expense not found", status_code=404)

    # 2. Must be an AI-extracted expense
    if expense.input_method != "ai":
        raise ExpenseServiceError(
            "Only AI-extracted expenses can be confirmed", status_code=409
        )

    # 3. Already confirmed
    if expense.is_confirmed:
        raise ExpenseServiceError("Expense is already confirmed", status_code=409)

    # 4. category_id must be set
    if expense.category_id is None:
        raise ExpenseServiceError(
            "Expense category is required before confirmation", status_code=422
        )

    # 5. Category must exist, be active, and not be soft-deleted
    category = (
        db.query(Category)
        .filter(
            Category.id == expense.category_id,
            Category.is_active == True,
            Category.deleted_at.is_(None),
        )
        .first()
    )
    if category is None:
        raise ExpenseServiceError(
            "A valid expense category is required before confirmation", status_code=422
        )

    # 6. Title must be present and not blank
    if not expense.title or not expense.title.strip():
        raise ExpenseServiceError(
            "Expense title is required before confirmation", status_code=422
        )

    # 7. total_amount must be set and zero or greater
    if expense.total_amount is None:
        raise ExpenseServiceError(
            "Expense total amount is required before confirmation", status_code=422
        )
    if expense.total_amount < 0:
        raise ExpenseServiceError(
            "Expense total amount must be zero or greater", status_code=422
        )

    # All checks passed — confirm in one transaction
    try:
        expense.is_confirmed = True
        db.commit()
        db.refresh(expense)
        items = _get_non_deleted_items(db, expense.id)
        return _build_expense_response(expense, items)
    except Exception:
        db.rollback()
        raise
